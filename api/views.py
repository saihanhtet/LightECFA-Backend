from django.core.mail import EmailMultiAlternatives, get_connection
import requests
from email import encoders
from email.mime.base import MIMEBase
from django.utils.timezone import now
from django.core.mail import EmailMultiAlternatives
import secrets
import string
from django.utils.timezone import make_aware
from datetime import datetime, date
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.generics import RetrieveAPIView
import json
from django.core.mail import send_mail
from django.contrib.auth.models import Group
from django.shortcuts import get_object_or_404
from rest_framework.generics import RetrieveUpdateDestroyAPIView
from django.db.models import Count
from django.contrib.auth import get_user_model, login, logout
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from rest_framework_simplejwt.authentication import JWTAuthentication as BaseJWTAuthentication
from rest_framework_simplejwt.tokens import RefreshToken

from api.models import *
from .serializers import *

import os
from dotenv import load_dotenv, find_dotenv

load_dotenv(find_dotenv())
UserModel = get_user_model()


class JWTAuthentication(BaseJWTAuthentication):
    """
    Custom JWT authentication class.
    This class extends the base JWTAuthentication provided by rest_framework_simplejwt.
    It adds support for checking access tokens stored in cookies if not found in the Authorization header.
    """

    def authenticate(self, request):
        # Check the Authorization header first
        header = self.get_header(request)
        if header is not None:
            raw_token = self.get_raw_token(header)
            validated_token = self.get_validated_token(raw_token)
            user = self.get_user(validated_token)
            return user, validated_token

        # If no token in the Authorization header, check cookies
        else:
            raw_token = request.COOKIES.get('access_token')
            if raw_token is not None:
                validated_token = self.get_validated_token(raw_token)
                user = self.get_user(validated_token)
                return user, validated_token

        return None


def generate_tokens(user):
    """ Generate access and refresh tokens for a given user. """
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


def EmailSender(subject, message, from_email, recipient_list, auth_user, auth_password):
    return send_mail(
        subject,
        message,
        from_email=from_email,
        recipient_list=recipient_list,
        auth_user=auth_user,
        auth_password=auth_password,
        fail_silently=False,
    )


def ResponseFunction(data, message, status, **args):
    """ Default reponse function layout """
    return Response({'data': data, **args, 'message': str(message)}, status=status)


def addCookies(response: Response, tokens):
    """ 
    Adding the cookies to our response header
    """
    response.set_cookie(
        key='refresh_token',
        value=tokens['refresh'],
        httponly=os.getenv('HTTPONLY'),
        secure=True,
        samesite='None',
        domain=os.getenv('DOMAIN'),
        path='/',
    )
    response.set_cookie(
        key='access_token',
        value=tokens['access'],
        httponly=os.getenv('HTTPONLY'),
        secure=True,
        samesite='None',
        domain=os.getenv('DOMAIN'),
        path='/',
    )
    return response


def GeneratePassword():
    characters = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(secrets.choice(characters) for _ in range(10))
    return password


def checkUserExist(email):
    try:
        user_obj = UserModel.objects.get(email=email)
        print(user_obj)
        return True
    except Exception as e:
        return False


class IsSuperuser(permissions.BasePermission):
    def has_permission(self, request, view):
        """ Check if the user making the request is a superuser"""
        return request.user and request.user.is_superuser


class IsTeacher(permissions.BasePermission):
    def has_permission(self, request, view):
        """ Check if the user making the request is a staff member"""
        return request.user and request.user.is_staff


class CheckToken(APIView):
    """ Check access token from a user """
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        """Check the token from request"""
        return Response({'data': {}, 'message': 'token alive'}, status=status.HTTP_200_OK)


class InvoiceFileUpload(APIView):
    serializer_class = InvoiceSerializer
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        invoices = Invoice.objects.all()
        serializer = self.serializer_class(invoices, many=True)
        return Response({'data': serializer.data, 'message': "successfully got the invoice file"}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"data": serializer.data, "message": "successfully added the invoice file"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ScheduledInvoiceMailView(APIView):
    serializer_class = ScheduledInvoiceMailSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            instance = EmailCredentials.objects.first()
            email_serializer = EmailCredentialsSerializer(instance)

        except EmailCredentials.DoesNotExist:
            return Response({'error': "Not found the object"}, status=status.HTTP_400_BAD_REQUEST)

        mails = ScheduledInvoiceMail.objects.all()
        serializers = self.serializer_class(mails, many=True)
        return Response({'data': serializers.data, "mail": email_serializer.data['email'], 'message': "successfully got the scheduled invoice mail"}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"data": serializer.data, "message": "successfully added a new invoice schedule"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ScheduledInvoiceMailDetail(RetrieveUpdateDestroyAPIView):
    queryset = ScheduledInvoiceMail.objects.all()
    serializer_class = ScheduledInvoiceMailSerializer
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]


class ScheduledInvoiceMailSend(RetrieveUpdateDestroyAPIView):
    queryset = ScheduledInvoiceMail.objects.all()
    serializer_class = ScheduledInvoiceMailSerializer
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]

    def post(self, request, pk):
        instance = self.get_object()
        invoice_serializer = self.serializer_class(instance).data

        data = EmailCredentials.objects.first()
        serializer = EmailCredentialsSerializer(data).data
        auth_user = serializer['email']
        auth_password = serializer['password']
        subject = invoice_serializer.get('subject')
        message = invoice_serializer.get('message')
        students = CustomUser.objects.filter(user_type='student')
        students = students.filter(
            studentprofile__cohort__course_id=invoice_serializer.get('course'))
        recipient_list = []
        attachment_response = requests.get(
            invoice_serializer.get('attachment'))

        for student in students:
            recipient_list.append(student.email)

        if len(recipient_list) == 0:
            return Response({'message': 'No Student for that course'}, status=status.HTTP_200_OK)
        print(recipient_list)
        if attachment_response.status_code == 200:
            email_message = EmailMultiAlternatives(
                subject=subject,
                body=message,
                from_email=auth_user,
                to=recipient_list,
            )

            attachment_content = attachment_response.content
            attachment_name = 'invoice.pdf'
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_content)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                            f'attachment; filename="{attachment_name}"')

            email_message.attach(part)

            email_message.connection = get_connection(
                username=auth_user,
                password=auth_password,
                fail_silently=False,
            )

            try:
                email_message.send(fail_silently=False)
                return Response({'message': 'Email sent successfully'}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'message': 'Failed to fetch attachment from URL:'}, status=status.HTTP_200_OK)


class InvoiceFileUploadDetail(RetrieveUpdateDestroyAPIView):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]


class EmailCredentialsDetailView(APIView):
    serializer_class = EmailCredentialsSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            instance = EmailCredentials.objects.first()
            serializer = EmailCredentialsSerializer(instance)
            return Response({"data": serializer.data, "message": "successfully got the email credentials"}, status=status.HTTP_200_OK)
        except EmailCredentials.DoesNotExist:
            return Response({'error': "Not found the object"}, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request):
        try:
            object = EmailCredentials.objects.first()
        except EmailCredentials.DoesNotExist:
            return Response({'error': "Not found the object"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.serializer_class(
            instance=object, data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"data": serializer.data, "message": "Email credentials updated successfully"}, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ======================== Register & Login View


class StudentRegisterView(APIView):
    """ View for creating students """
    serializer_class = StudentRegisterSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        serializer = self.serializer_class()
        return Response(serializer.data)

    def post(self, request):
        clean_data = request.data.copy()
        clean_data['password'] = str(GeneratePassword())
        cohort_id = int(clean_data['cohort'][0])
        try:
            cohort = Cohort.objects.get(pk=cohort_id)
            course = cohort.course
        except Cohort.DoesNotExist:
            pass
        exists = checkUserExist(clean_data['email'])
        if exists:
            return Response({'data': {}, 'message': 'User with this Email Exist'}, status=status.HTTP_226_IM_USED)
        serializer = StudentRegisterSerializer(data=clean_data)
        if serializer.is_valid(raise_exception=True):
            user = serializer.create(clean_data)
            enrolled_course = course
            subject = "Welcome to Light English Class For All!"
            user = UserModel.objects.get(id=user.id)
            serializer = UserDetailSerializer(user)
            email_data = EmailCredentials.objects.first()
            email_serializer = EmailCredentialsSerializer(email_data).data
            from_email = email_serializer['email']
            auth_password = email_serializer['password']
            subject = 'Hello'
            message = 'Here is the testing from django.'
            recipient_list = [serializer.data['email']]
            if user:
                EmailSender(subject, message, from_email,
                            recipient_list, from_email, auth_password)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response({"data": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        return Response({"data": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class TeacherRegisterView(APIView):
    """ View for creating teacher """
    serializer_class = TeacherRegisterSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        serializer = self.serializer_class()
        return Response(serializer.data)

    def post(self, request):
        clean_data = request.data.copy()
        clean_data['password'] = str(GeneratePassword())
        exists = checkUserExist(clean_data['email'])
        if exists:
            return Response({'data': {}, 'message': 'User with this Email Exist'}, status=status.HTTP_226_IM_USED)
        serializer = TeacherRegisterSerializer(data=clean_data)
        if serializer.is_valid(raise_exception=True):
            user = serializer.create(clean_data)
            user = UserModel.objects.get(id=user.id)
            serializer = UserDetailSerializer(user)
            email_data = EmailCredentials.objects.first()
            email_serializer = EmailCredentialsSerializer(email_data).data
            from_email = email_serializer['email']
            auth_password = email_serializer['password']
            subject = 'Welcome from Light English Class For All'
            message = f"Here is your credentials to login: \n Email: {clean_data['email']} ,\n Password: {clean_data['password']}"
            recipient_list = [serializer.data['email']]
            if user:
                EmailSender(subject, message, from_email,
                            recipient_list, from_email, auth_password)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class AdminRegisterView(APIView):
    """ View for creating admin """
    serializer_class = AdminRegisterSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        serializer = self.serializer_class()
        return Response(serializer.data)

    def post(self, request):
        clean_data = request.data.copy()
        clean_data['password'] = str(GeneratePassword())
        exists = checkUserExist(clean_data['email'])
        if exists:
            return Response({'data': {}, 'message': 'User with this Email Exist'}, status=status.HTTP_226_IM_USED)
        serializer = AdminRegisterSerializer(data=clean_data)
        if serializer.is_valid(raise_exception=True):
            user = serializer.create(clean_data)
            user = UserModel.objects.get(id=user.id)
            serializer = UserDetailSerializer(user)
            email_data = EmailCredentials.objects.first()
            email_serializer = EmailCredentialsSerializer(email_data).data
            from_email = email_serializer['email']
            auth_password = email_serializer['password']
            subject = 'Welcome from Light English Class For All'
            message = f"Here is your credentials to login: \n Email: {clean_data['email']} ,\n Password: {clean_data['password']}"
            recipient_list = [serializer.data['email']]
            if user:
                EmailSender(subject, message, from_email,
                            recipient_list, from_email, auth_password)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class RegisterView(APIView):
    """
    View for registering user.
    """
    serializer_class = UserRegisterSerializer
    permission_classes = (permissions.AllowAny,)

    def post(self, request):
        """ User Register post method """
        clean_data = request.data
        serializer = UserRegisterSerializer(data=clean_data)
        if serializer.is_valid(raise_exception=True):
            user = serializer.create(clean_data)
            serializer = UserRegisterSerializer(user)
            if user:
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    """ View for user login. """
    serializer_class = UserLoginSerializer
    permission_classes = (permissions.AllowAny,)

    def post(self, request):
        """ User Login post method """
        data = request.data
        if data['email'] == "":
            return Response({"data": "", "message": "Please fill the email"}, status=status.HTTP_100_CONTINUE)
        if data['password'] == "":
            return Response({"data": "", "message": "Please fill the password"}, status=status.HTTP_100_CONTINUE)

        serializer = UserLoginSerializer(data=data)
        if serializer.is_valid(raise_exception=True):
            user = serializer.check_user(data)
            login(request, user)
            tokens = generate_tokens(user)
            serializer = UserDetailSerializer(user)
            response = ResponseFunction(
                data=serializer.data, message="Successfully Login", status=status.HTTP_200_OK, **tokens)

            response = addCookies(response, tokens)
            return response


class LogoutView(APIView):
    permission_classes = (permissions.IsAuthenticated, )
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        """ User Logout method """
        logout(request)
        response = ResponseFunction(
            {}, "Successfully Logout", status.HTTP_200_OK)
        response.delete_cookie('refresh_token')
        response.delete_cookie('access_token')
        return response

# ======================== Profile View


class TeacherProfileDetailView(RetrieveUpdateDestroyAPIView):
    queryset = TeacherProfile.objects.all()
    serializer_class = TeacherProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]


class StudentProfileDetailView(RetrieveUpdateDestroyAPIView):
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        user_id = self.kwargs['pk']
        queryset = StudentProfile.objects.filter(user=user_id)
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if queryset.exists():
            instance = queryset.first()
            serializer = self.serializer_class(instance)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({"detail": "StudentProfile not found for the given user ID."}, status=status.HTTP_404_NOT_FOUND)


class AdminProfileDetailView(RetrieveUpdateDestroyAPIView):
    queryset = AdminProfile.objects.all()
    serializer_class = AdminProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]


class UserTypeChangeView(APIView):
    serializer_class = UserTypeChangeSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        serializer = UserTypeChangeSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': "Successfully updated the user type"}, status=status.HTTP_200_OK)
        return Response({'data': "", 'message': "Check your password"}, status=status.HTTP_100_CONTINUE)


class UserDetailView(RetrieveUpdateDestroyAPIView):
    """
    View for getting, updating, and deleting user details.
    """
    queryset = UserModel.objects.all()
    serializer_class = UserDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        """ User Details get method """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        message = "Successfully retrieved user details."
        return Response({"data": serializer.data, "message": message}, status=status.HTTP_200_OK)

    def put(self, request, *args, **kwargs):
        """ User Details put method """
        instance = self.get_object()
        new_user_type = request.data.get('user_type')
        previous_user_type = instance.user_type
        data = {}
        if new_user_type != previous_user_type:
            try:
                # Delete previous profile if it exists
                if previous_user_type == CustomUser.ADMIN:
                    try:
                        admin_profile = AdminProfile.objects.get(user=instance)
                        data = {
                            'sex': admin_profile.sex,
                            'phone_number': admin_profile.phone_number,
                            'date_of_birth': admin_profile.date_of_birth,
                            'profile_pic': admin_profile.profile_pic,
                            'address': admin_profile.address,
                        }
                        admin_profile.delete()
                    except AdminProfile.DoesNotExist:
                        pass
                elif previous_user_type == CustomUser.TEACHER:
                    try:
                        teacher_profile = TeacherProfile.objects.get(
                            user=instance)
                        data = {
                            'sex': teacher_profile.sex,
                            'phone_number': teacher_profile.phone_number,
                            'date_of_birth': teacher_profile.date_of_birth,
                            'profile_pic': teacher_profile.profile_pic,
                            'address': teacher_profile.address,
                        }
                        teacher_profile.delete()
                    except TeacherProfile.DoesNotExist:
                        pass
                elif previous_user_type == CustomUser.STUDENT:
                    try:
                        student_profile = StudentProfile.objects.get(
                            user=instance)
                        data = {
                            'sex': student_profile.sex,
                            'phone_number': student_profile.phone_number,
                            'date_of_birth': student_profile.date_of_birth,
                            'profile_pic': student_profile.profile_pic,
                            'address': student_profile.address,
                        }
                        student_profile.delete()
                    except StudentProfile.DoesNotExist:
                        pass
            except Exception as e:
                print(f"Error deleting previous profile: {e}")

            try:
                # Create new profile based on new user type
                if new_user_type == CustomUser.ADMIN:
                    AdminProfile.objects.create(user=instance, **data)
                    instance.is_superuser = True
                    instance.is_staff = True
                    admin_group = Group.objects.get(name='Admins')
                    instance.groups.clear()
                    instance.groups.add(admin_group)
                elif new_user_type == CustomUser.TEACHER:
                    TeacherProfile.objects.create(user=instance, **data)
                    instance.is_superuser = False
                    instance.is_staff = True
                    teacher_group = Group.objects.get(name='Teachers')
                    instance.groups.clear()
                    instance.groups.add(teacher_group)
                elif new_user_type == CustomUser.STUDENT:
                    StudentProfile.objects.create(user=instance, **data)
                    instance.is_superuser = False
                    instance.is_staff = False
                    student_group = Group.objects.get(name='Students')
                    instance.groups.clear()
                    instance.groups.add(student_group)
            except Exception as e:
                print(f"Error creating new profile: {e}")

        serializer = self.get_serializer(instance, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        """ User Details delete method """
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class UserListView(APIView):
    """
    View for retrieving user data.
    """
    serializer_class = UserDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, user_type=None):
        if user_type:
            all_user = UserModel.objects.filter(
                user_type=user_type).order_by('id')
            message = f"Successfully retrieve {user_type} users"
        else:
            all_user = UserModel.objects.all().order_by('id')
            message = "Successfully retrieve all users"

        serializer = self.serializer_class(all_user, many=True)
        return Response({'data': serializer.data, 'message': message}, status=status.HTTP_200_OK)

# ======================== Others View


class SubjectView(APIView):
    serializer_class = SubjectDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        subjects = Subject.objects.all()
        serializer = self.serializer_class(subjects, many=True)

        if subjects.count() == 0:
            return Response({'data': serializer.data, 'message': "No subjects."}, status=status.HTTP_200_OK)
        return Response({'data': serializer.data, 'message': "Successfully retrieved Subject."}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': 'Successfully inserted new subject'}, status=status.HTTP_201_CREATED)
        return Response({'data': "", "message": "Please fill all the input"}, status=status.HTTP_400_BAD_REQUEST)


class SubjectDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)


class CourseView(APIView):
    serializer_class = CourseDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        courses = Course.objects.all()
        serializer = self.serializer_class(courses, many=True)

        if courses.count() == 0:
            return Response({'data': serializer.data, 'message': "No courses."}, status=status.HTTP_200_OK)
        return Response({'data': serializer.data, 'message': "Successfully retrieved courses."}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': 'Successfully inserted new course'}, status=status.HTTP_201_CREATED)
        return Response({'data': "", "message": "Please fill all the input"}, status=status.HTTP_400_BAD_REQUEST)


class CourseDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)


class CohortView(APIView):
    serializer_class = CohortDetailSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        cohorts = Cohort.objects.all()
        serializer = self.serializer_class(cohorts, many=True)

        if cohorts.count() == 0:
            return Response({'data': serializer.data, 'message': "No cohorts."}, status=status.HTTP_200_OK)
        return Response({'data': serializer.data, 'message': "Successfully retrieved Cohort."}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': 'Successfully inserted new cohort'}, status=status.HTTP_201_CREATED)
        return Response({'data': "", "message": "Please fill all the input"}, status=status.HTTP_400_BAD_REQUEST)


class CohortDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Cohort.objects.all()
    serializer_class = CohortDetailSerializer
    permission_classes = (permissions.IsAuthenticated, )


class AttendanceView(APIView):
    serializer_class = AttendanceSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        return Response({'data': "", "message": "Successfully got the attendance"}, status=status.HTTP_200_OK)


class Dashboard(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        users = UserModel.objects.all()
        current_user = UserModel.objects.get(pk=int(request.user.id))
        current_user_serializer = UserDetailSerializer(current_user).data

        graduate_student = StudentProfile.objects.filter(
            status=StudentProfile.GRADUATE)
        undergraduate_student = StudentProfile.objects.filter(
            status=StudentProfile.UNDERGRADUATE)

        admins = UserModel.objects.filter(user_type='admin')
        teachers = UserModel.objects.filter(user_type='teacher')
        students = UserModel.objects.filter(user_type='student')

        users_serializer = UserDetailSerializer(users, many=True).data
        admins_serializer = UserDetailSerializer(admins, many=True).data
        teachers_serializer = UserDetailSerializer(teachers, many=True).data
        students_serializer = UserDetailSerializer(students, many=True).data

        graduate_students = [UserModel.objects.get(
            pk=user.user.id) for user in graduate_student]

        undergraduate_students = [UserModel.objects.get(
            pk=user.user.id) for user in undergraduate_student]

        graduate_student_serializer = UserDetailSerializer(
            graduate_students, many=True).data
        undergraduate_student_serializer = UserDetailSerializer(
            undergraduate_students, many=True).data

        courses = Course.objects.all()
        subjects = Subject.objects.all()
        cohorts = Cohort.objects.all()
        invoice = Invoice.objects.all()
        tests = TestModel.objects.all()
        test_reports = TestStudentReport.objects.all()

        courses_serializer = CourseDetailSerializer(courses, many=True).data
        subjects_serializer = SubjectDetailSerializer(subjects, many=True).data
        cohorts_serializer = CohortDetailSerializer(cohorts, many=True).data
        invoice_serializer = InvoiceSerializer(invoice, many=True).data

        tests_serializer = TestModelSerializer(tests, many=True).data
        test_reports_serializer = TestStudentReportSerilizer(
            test_reports, many=True).data

        invoice_mails = ScheduledInvoiceMail.objects.all()
        invoice_serializers = ScheduledInvoiceMailSerializer(
            invoice_mails, many=True).data

        email_instance = EmailCredentials.objects.all()
        email_serializer = EmailCredentialsSerializer(
            email_instance, many=True).data

        data = {
            "current_user": current_user_serializer,
            "users": users_serializer,
            "admins": admins_serializer,
            "teachers": teachers_serializer,
            "students": students_serializer,
            "invoices": invoice_serializer,
            "undergraduate_students": undergraduate_student_serializer,
            "graduate_students": graduate_student_serializer,
            "users_count": users.count(),
            "admins_count": admins.count(),
            "teachers_count": teachers.count(),
            "students_count": students.count(),
            "undergraduate_count": undergraduate_student.count(),
            "graduate_count": graduate_student.count(),
            "courses": courses_serializer,
            "subjects": subjects_serializer,
            "cohorts": cohorts_serializer,
            "courses_count": courses.count(),
            "subjects_count": subjects.count(),
            "cohorts_count": cohorts.count(),
            "invoice_schedule": invoice_serializers,
            "tests": tests_serializer,
            "test_reports": test_reports_serializer,
            "default_mail": email_serializer[0],
        }

        return Response({'data': data, 'message': "Successfully got the all information."}, status=status.HTTP_200_OK)


class SendAllInvoice(APIView):
    serializer_class = InvoiceSerializer
    permission_classes = (permissions.IsAuthenticated, )
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        students = UserModel.objects.filter(user_type="student")
        emails = [student.email for student in students]
        print(emails)
        return Response({'data': "", "message": "Successfully sent the invoices to all student"}, status=status.HTTP_200_OK)


class SendEmailView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        data = EmailCredentials.objects.first()
        serializer = EmailCredentialsSerializer(data).data
        email = serializer['email']
        password = serializer['password']
        subject = 'Hello'
        message = 'Here is the testing from django.'
        recipient_list = ['hanhtetsan13@gmail.com']
        try:
            send_mail(
                subject,
                message,
                from_email=email,
                recipient_list=recipient_list,
                auth_user=email,
                auth_password=password,
                fail_silently=False,
            )
            return Response({'message': 'Email sent successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class TestMakerView(APIView):
    serializer_class = TestModelSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )

    def get(self, request):
        test = TestModel.objects.all()
        serializer = self.serializer_class(test, many=True)
        if not test.exists():
            return Response({'data': serializer.data, 'message': "No test."}, status=status.HTTP_200_OK)
        return Response({'data': serializer.data, 'message': "Successfully retrieved Test."}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': 'Successfully inserted new test'}, status=status.HTTP_201_CREATED)
        return Response({'data': "", "message": "Please fill all the input"}, status=status.HTTP_400_BAD_REQUEST)


class TestMakerDetailView(RetrieveUpdateDestroyAPIView):
    queryset = TestModel.objects.all()
    serializer_class = TestModelSerializer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )


class TestReportView(APIView):
    serializer_class = TestStudentReportSerilizer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )

    def get(self, request):
        test_report = TestStudentReport.objects.all()
        serializer = self.serializer_class(test_report, many=True)

        if not test_report.exists():
            return Response({'data': serializer.data, 'message': "No test report."}, status=status.HTTP_200_OK)
        return Response({'data': serializer.data, 'message': "Successfully retrieved Test."}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({'data': serializer.data, 'message': 'Successfully inserted new test report'}, status=status.HTTP_201_CREATED)
        return Response({'data': "", "message": "Please fill all the input"}, status=status.HTTP_400_BAD_REQUEST)


class TestReportDetailView(RetrieveUpdateDestroyAPIView):
    queryset = TestStudentReport.objects.all()
    serializer_class = TestStudentReportSerilizer
    permission_classes = (permissions.IsAuthenticated, IsSuperuser, )
# excels


class AdminUserExcelView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSuperuser,)
    authentication_classes = [JWTAuthentication]

    def get_age(self, date_of_birth):
        today = date.today()
        return today.year - date_of_birth.year - ((today.month, today.day) < (date_of_birth.month, date_of_birth.day))

    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        filename = "lightecfa_admins.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Admins"

        headers = ["Id", "Name", "Email", "Age",
                   "Sex", 'Phone Number', 'Address', "Picture"]
        ws.append(headers)

        admins = UserModel.objects.filter(user_type='admin')
        admins_data = UserDetailSerializer(admins, many=True).data

        for admin in admins_data:
            profile = admin.get('profile')
            date_of_birth = profile.get('date_of_birth')
            if date_of_birth:
                date_of_birth = datetime.strptime(
                    date_of_birth, '%Y-%m-%d').date()
                age = self.get_age(date_of_birth)
            else:
                age = 'N/A'
            ws.append([
                profile.get('id'),
                admin.get('username'),
                admin.get('email'),
                age,
                profile.get('sex'),
                profile.get('phone_number'),
                profile.get('address'),
                profile.get('profile_pic'),
            ])

        wb.save(response)
        return response


class StudentUserExcelView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSuperuser,)
    authentication_classes = [JWTAuthentication]

    def get_age(self, date_of_birth):
        today = date.today()
        return today.year - date_of_birth.year - ((today.month, today.day) < (date_of_birth.month, date_of_birth.day))

    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        filename = "lightecfa_students.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = ["Student ID", "Name", "Email", "Age",
                   "Sex", 'Phone Number', 'Address', "Picture",
                   'Guardian Name', 'Guardian Phone', 'Guardian Phone 2', 'Status', 'Cohort', 'Subjects']
        ws.append(headers)

        students = UserModel.objects.filter(user_type='student')
        students_data = UserDetailSerializer(students, many=True).data

        for student in students_data:
            profile = student.get('profile')
            date_of_birth = profile.get('date_of_birth')
            date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            age = self.get_age(date_of_birth)
            try:
                cohort = Cohort.objects.get(
                    pk=profile.get('cohort')).cohort_name
            except Cohort.DoesNotExist:
                cohort = 'N/A'
            subjects = []
            if len(list(profile.get('subject'))) > 1:
                for subjId in list(profile.get('subject')):
                    try:
                        subject = Subject.objects.get(pk=subjId).subject_name
                        subjects.append(subject)
                    except Subject.DoesNotExist:
                        continue
            subjects = ",".join(subjects)
            ws.append([
                profile.get('student_id'),
                student.get('username'),
                student.get('email'),
                age,
                profile.get('sex'),
                profile.get('phone_number'),
                profile.get('address'),
                profile.get('profile_pic'),
                profile.get('guardian_name'),
                profile.get('guardian_phone'),
                profile.get('guardian_phone2'),
                profile.get('status'),
                cohort,
                subjects,
            ])

        wb.save(response)
        return response


class TeacherUserExcelView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSuperuser,)
    authentication_classes = [JWTAuthentication]

    def get_age(self, date_of_birth):
        today = date.today()
        return today.year - date_of_birth.year - ((today.month, today.day) < (date_of_birth.month, date_of_birth.day))

    def get(self, request):
        response = HttpResponse(
            content_type='application/ms-excel')
        filename = "lightecfa_teachers.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = ["ID", "Name", "Email", "Age",
                   "Sex", 'Phone Number', 'Address', "Picture", 'Subjects']
        ws.append(headers)

        students = UserModel.objects.filter(user_type='teacher')
        students_data = UserDetailSerializer(students, many=True).data

        for teacher in students_data:
            profile = teacher.get('profile')
            date_of_birth = profile.get('date_of_birth')
            date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            age = self.get_age(date_of_birth)
            subjects = []
            if len(list(profile.get('subject'))) > 1:
                for subjId in list(profile.get('subject')):
                    try:
                        subject = Subject.objects.get(pk=subjId).subject_name
                        subjects.append(subject)
                    except Subject.DoesNotExist:
                        continue
            subjects = ",".join(subjects)
            ws.append([
                teacher.get('id'),
                teacher.get('username'),
                teacher.get('email'),
                age,
                profile.get('sex'),
                profile.get('phone_number'),
                profile.get('address'),
                profile.get('profile_pic'),
                subjects,
            ])

        wb.save(response)
        return response
